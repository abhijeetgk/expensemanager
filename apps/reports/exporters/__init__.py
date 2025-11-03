"""
Export services using Strategy pattern.

Demonstrates:
- Strategy pattern for different export formats
- Abstract base classes
- Type hints
- File handling
- Context managers
"""

from abc import ABC, abstractmethod
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from apps.reports.services import ReportService, CategorySummary


class ExporterBase(ABC):
    """
    Abstract base class for export strategies.
    
    This demonstrates the Strategy pattern where different
    export formats implement this interface.
    """
    
    @abstractmethod
    def export(self, data: Dict[str, Any], filename: str) -> BytesIO:
        """
        Export data to a specific format.
        
        Args:
            data: Data to export
            filename: Name for the exported file
            
        Returns:
            BytesIO object containing the exported data
        """
        pass
    
    @abstractmethod
    def get_content_type(self) -> str:
        """
        Get the MIME type for this export format.
        
        Returns:
            Content type string
        """
        pass
    
    @abstractmethod
    def get_file_extension(self) -> str:
        """
        Get the file extension for this export format.
        
        Returns:
            File extension string (with dot)
        """
        pass


class ExcelExporter(ExporterBase):
    """
    Excel export strategy using openpyxl.
    
    Demonstrates the Strategy pattern implementation for Excel exports.
    """
    
    def get_content_type(self) -> str:
        """Return Excel MIME type."""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def get_file_extension(self) -> str:
        """Return Excel file extension."""
        return '.xlsx'
    
    def export(self, data: Dict[str, Any], filename: str) -> BytesIO:
        """
        Export data to Excel format.
        
        Args:
            data: Data to export
            filename: Name for the exported file
            
        Returns:
            BytesIO object containing the Excel file
        """
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        # Create summary sheet
        self._create_summary_sheet(workbook, data.get('summary', {}))
        
        # Create category breakdown sheet
        if 'category_breakdown' in data:
            self._create_category_sheet(workbook, data['category_breakdown'])
        
        # Create transactions sheet
        if 'transactions' in data:
            self._create_transactions_sheet(workbook, data['transactions'])
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, summary: Dict[str, Any]) -> None:
        """Create summary sheet with styling."""
        ws = workbook.create_sheet("Summary", 0)
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "Expense Manager Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Period
        ws['A2'] = f"Period: {summary.get('period_start', '')} to {summary.get('period_end', '')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        # Summary data
        row = 4
        summary_items = [
            ("Total Income", summary.get('total_income', 0)),
            ("Total Expense", summary.get('total_expense', 0)),
            ("Net Balance", summary.get('net_balance', 0)),
            ("Transaction Count", summary.get('transaction_count', 0)),
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].number_format = '#,##0.00' if isinstance(value, (int, float, Decimal)) else None
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_category_sheet(self, workbook: openpyxl.Workbook, categories: List[Dict[str, Any]]) -> None:
        """Create category breakdown sheet."""
        ws = workbook.create_sheet("Category Breakdown")
        
        # Headers
        headers = ["Category", "Amount", "Count", "Percentage"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for category in categories:
            ws.append([
                category.get('category_name', ''),
                category.get('total_amount', 0),
                category.get('transaction_count', 0),
                f"{category.get('percentage', 0):.2f}%"
            ])
        
        # Format amount column
        for row in range(2, len(categories) + 2):
            ws[f'B{row}'].number_format = '#,##0.00'
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    def _create_transactions_sheet(self, workbook: openpyxl.Workbook, transactions: List[Dict[str, Any]]) -> None:
        """Create transactions detail sheet."""
        ws = workbook.create_sheet("Transactions")
        
        # Headers
        headers = ["Date", "Description", "Category", "Amount", "Type"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for transaction in transactions:
            ws.append([
                transaction.get('date', ''),
                transaction.get('description', ''),
                transaction.get('category', ''),
                transaction.get('amount', 0),
                transaction.get('type', '')
            ])
        
        # Format amount column
        for row in range(2, len(transactions) + 2):
            ws[f'D{row}'].number_format = '#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10


class PDFExporter(ExporterBase):
    """
    PDF export strategy using ReportLab.
    
    Demonstrates the Strategy pattern implementation for PDF exports.
    """
    
    def get_content_type(self) -> str:
        """Return PDF MIME type."""
        return 'application/pdf'
    
    def get_file_extension(self) -> str:
        """Return PDF file extension."""
        return '.pdf'
    
    def export(self, data: Dict[str, Any], filename: str) -> BytesIO:
        """
        Export data to PDF format.
        
        Args:
            data: Data to export
            filename: Name for the exported file
            
        Returns:
            BytesIO object containing the PDF file
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        story.append(Paragraph("Expense Manager Report", title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Summary section
        if 'summary' in data:
            self._add_summary_section(story, data['summary'], styles)
        
        story.append(Spacer(1, 0.3 * inch))
        
        # Category breakdown
        if 'category_breakdown' in data:
            self._add_category_section(story, data['category_breakdown'], styles)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        return output
    
    def _add_summary_section(self, story: List, summary: Dict[str, Any], styles) -> None:
        """Add summary section to PDF."""
        # Section title
        story.append(Paragraph("Summary", styles['Heading2']))
        story.append(Spacer(1, 0.1 * inch))
        
        # Period
        period_text = f"Period: {summary.get('period_start', '')} to {summary.get('period_end', '')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 0.2 * inch))
        
        # Summary table
        data = [
            ['Metric', 'Value'],
            ['Total Income', f"₹{summary.get('total_income', 0):,.2f}"],
            ['Total Expense', f"₹{summary.get('total_expense', 0):,.2f}"],
            ['Net Balance', f"₹{summary.get('net_balance', 0):,.2f}"],
            ['Transaction Count', str(summary.get('transaction_count', 0))],
        ]
        
        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(table)
    
    def _add_category_section(self, story: List, categories: List[Dict[str, Any]], styles) -> None:
        """Add category breakdown section to PDF."""
        # Section title
        story.append(Paragraph("Category Breakdown", styles['Heading2']))
        story.append(Spacer(1, 0.1 * inch))
        
        # Category table
        data = [['Category', 'Amount', 'Count', 'Percentage']]
        
        for category in categories:
            data.append([
                category.get('category_name', ''),
                f"₹{category.get('total_amount', 0):,.2f}",
                str(category.get('transaction_count', 0)),
                f"{category.get('percentage', 0):.2f}%"
            ])
        
        table = Table(data, colWidths=[2.5 * inch, 1.5 * inch, 1 * inch, 1 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(table)


class ExportService:
    """
    Service for managing exports using different strategies.
    
    This demonstrates the Context class in the Strategy pattern.
    """
    
    def __init__(self, exporter: ExporterBase):
        """
        Initialize with a specific export strategy.
        
        Args:
            exporter: The export strategy to use
        """
        self._exporter = exporter
    
    def export_report(
        self,
        user,
        report_type: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> tuple[BytesIO, str, str]:
        """
        Export a report using the configured strategy.
        
        Args:
            user: The user to export for
            report_type: Type of report ('summary', 'category', etc.)
            start_date: Optional start date
            end_date: Optional end date
            
        Returns:
            Tuple of (file_content, content_type, filename)
        """
        # Gather report data
        summary = ReportService.get_summary_report(user, start_date, end_date)
        
        data = {
            'summary': {
                'total_income': summary.total_income,
                'total_expense': summary.total_expense,
                'net_balance': summary.net_balance,
                'transaction_count': summary.transaction_count,
                'period_start': summary.period_start,
                'period_end': summary.period_end,
            }
        }
        
        # Add category breakdown
        expense_categories = ReportService.get_category_wise_report(
            user, 'EXPENSE', start_date, end_date
        )
        data['category_breakdown'] = [
            {
                'category_name': cat.category_name,
                'total_amount': cat.total_amount,
                'transaction_count': cat.transaction_count,
                'percentage': cat.percentage,
            }
            for cat in expense_categories
        ]
        
        # Generate filename
        filename = f"expense_report_{summary.period_start}_{summary.period_end}"
        filename += self._exporter.get_file_extension()
        
        # Export using strategy
        file_content = self._exporter.export(data, filename)
        
        return file_content, self._exporter.get_content_type(), filename

